import time

import pyautogui
from openpyxl import workbook, load_workbook

class FileHelp:

    def reader_excel(self, filename, sheetname):
        wb = load_workbook(filename=filename)
        sh = wb[sheetname]
        datalist = []

        rownumber = sh.max_row
        columnumber = sh.max_column

        for i in range(1, rownumber + 1):
            row = []
            for j in range(1, columnumber + 1):
                row.append(sh.cell(row=i, column=j).value)
            datalist.append(row)

        return datalist

    '''
        upload file with pyautogui
    '''
    def upload_file(self, filepath):
        time.sleep(1)
        pyautogui.write(filepath)
        pyautogui.press("enter")